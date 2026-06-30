from __future__ import annotations

import json
from copy import deepcopy
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "_input" / "WO_MB_Tracker_May2026.xlsx"
OUTPUT_PATH = ROOT / "data" / "private" / "local-seed" / "openforge-tracker-seed.json"


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.2f}".rstrip("0").rstrip(".")
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value).strip()


def scale_money(value: str, multiplier: float) -> str:
    try:
        return f"{round(float(value) * multiplier, 2):.2f}"
    except ValueError:
        return value


def rows_to_dicts(sheet, row_numbers: list[int], keys: dict[str, str]) -> list[dict[str, str]]:
    headers = [text(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    header_index = {header: index for index, header in enumerate(headers)}
    records: list[dict[str, str]] = []
    for row_number in row_numbers:
        row = list(sheet.iter_rows(min_row=row_number, max_row=row_number, values_only=True))[0]
        records.append(
            {
                output_key: text(row[header_index[source_header]])
                for output_key, source_header in keys.items()
            }
        )
    return records


def build_profile_two_rows(rows: list[dict[str, str]], money_keys: list[str], prefix: str) -> list[dict[str, str]]:
    duplicated = deepcopy(rows)
    for row in duplicated:
        for money_key in money_keys:
            if money_key in row:
                row[money_key] = scale_money(row[money_key], 1.35)
        if "id" in row:
            row["id"] = row["id"].replace("IT1-", prefix)
    return duplicated


def main() -> None:
    workbook = load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)

    sportsbook = rows_to_dicts(
        workbook["Sportsbook Bets"],
        [2, 4, 5, 7, 8],
        {
            "id": "QualBetID",
            "dateSettling": "DateSettling",
            "eventName": "EventName",
            "offer": "Offer",
            "bookmaker": "Bookmaker",
            "offerType": "OfferType",
            "status": "Status",
            "result": "Result",
            "backStake": "BackStake",
            "backOdds": "BackOdds",
            "matchStrategy": "MatchStrategy",
            "layOdds1": "LayOdds1",
            "exchange": "Exchange",
        },
    )
    free_bets = rows_to_dicts(
        workbook["Free Bets"],
        [2, 4, 6, 7, 8],
        {
            "id": "FreeBetID",
            "dateSettling": "DateSettling",
            "eventName": "EventName",
            "offer": "Offer",
            "bookmaker": "Bookmaker",
            "status": "Status",
            "result": "Result",
            "retentionMode": "FreeBetRetentionMode",
            "freeBetValue": "FreeBetValue",
            "backOdds": "BackOdds",
            "matchStrategy": "MatchStrategy",
            "layOdds1": "LayOdds1",
            "exchange": "Exchange",
            "expiryDateTime": "ExpiryDateTime",
        },
    )
    casino_offers = rows_to_dicts(
        workbook["Casino Offers"],
        [2, 3, 5, 7, 8],
        {
            "id": "CasinoOfferID",
            "dateStarted": "DateStarted",
            "dateSettling": "DateSettling",
            "bookmaker": "Bookmaker",
            "offerType": "OfferType",
            "offerName": "OfferName",
            "game": "Game",
            "cashStake": "CashStake",
            "freeSpinsAwarded": "Free Spins Awarded",
            "freeSpinsValue": "Free Spins Value",
            "status": "Status",
            "result": "Result",
        },
    )
    accounts = rows_to_dicts(
        workbook["Accounts"],
        [2, 3, 4, 5, 6],
        {
            "id": "AccountID",
            "account": "Account",
            "type": "Type",
            "countsInCashTotal": "Counts In Cash Total",
            "channel": "Channel",
            "status": "Status",
            "currentBalance": "CurrentBalance",
            "pendingWithdrawalAmount": "PendingWithdrawalAmount",
            "lastBalanceUpdate": "LastBalanceUpdate",
            "group": "Group",
            "platform": "Platform",
        },
    )
    cash_adjustments = rows_to_dicts(
        workbook["Cash Adjustments"],
        [2, 3, 4, 5, 6],
        {
            "id": "AdjustmentID",
            "adjustmentDate": "AdjustmentDate",
            "direction": "Direction",
            "amount": "Amount",
            "adjustmentType": "AdjustmentType",
            "affectsInvestment": "AffectsInvestment",
            "affectsCashSnapshot": "AffectsCashSnapshot",
            "linkedAccount": "LinkedAccount",
            "description": "Description",
        },
    )

    profile_one = {
        "profileId": "profile-demo-001",
        "displayName": "Subscriber Alpha",
        "profileCode": "ALPHA-001",
        "status": "active",
        "trackingStartDate": "2026-05-01",
        "managementFeePercent": "40.00",
        "investmentFeePercent": "0.00",
        "currentCashSnapshot": "Workbook-derived local seed",
        "trackerData": {
            "accounts": accounts,
            "sportsbook-bets": sportsbook,
            "free-bets": free_bets,
            "casino-offers": casino_offers,
            "cash-adjustments": cash_adjustments,
        },
    }

    profile_two = {
        "profileId": "profile-demo-002",
        "displayName": "Subscriber Bravo",
        "profileCode": "BRAVO-002",
        "status": "paused",
        "trackingStartDate": "2026-05-15",
        "managementFeePercent": "35.00",
        "investmentFeePercent": "5.00",
        "currentCashSnapshot": "Workbook-derived local seed",
        "trackerData": {
            "accounts": build_profile_two_rows(accounts, ["currentBalance", "pendingWithdrawalAmount"], "IT2-"),
            "sportsbook-bets": build_profile_two_rows(sportsbook, ["backStake", "backOdds", "layOdds1"], "IT2-"),
            "free-bets": build_profile_two_rows(free_bets, ["freeBetValue", "backOdds", "layOdds1"], "IT2-"),
            "casino-offers": build_profile_two_rows(casino_offers, ["cashStake", "freeSpinsAwarded", "freeSpinsValue"], "IT2-"),
            "cash-adjustments": build_profile_two_rows(cash_adjustments, ["amount"], "IT2-"),
        },
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        json.dumps(
            {
                "generatedAt": datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z"),
                "sourceWorkbook": str(WORKBOOK_PATH.name),
                "profiles": [profile_one, profile_two],
            },
            indent=2,
        )
    )
    print(f"Wrote local seed to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
